# Запись результата жеребьёвки в .xlsx — два листа: участники и сетка.
from __future__ import annotations
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Participant


def write_excel(
    participants: List[Participant],
    assignment: Dict[Tuple[int, int, str], List[str]],
    output_path,
) -> None:
    # Первый лист — список участников, второй — собственно сетка. output_path может быть строкой-путём или file-like объектом (например, io.BytesIO).
    wb = Workbook()

    ws_p = wb.active
    ws_p.title = "Участники"
    ws_p.append(["#", "Имя", "Уровень", "Айрон", "Открывающий"])
    for i, p in enumerate(participants, start=1):
        ws_p.append([
            i, p.name, p.level,
            "да" if p.ironman else "нет",
            p.opening if p.ironman else "",
        ])

    ws = wb.create_sheet("Draw")

    # Палитра: синий для ПРОП (правительство), жёлтый для ОПП (оппозиция);
    # тёмные оттенки в шапке, светлые в теле.
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    gov_fill = PatternFill("solid", fgColor="5B88B4")
    opp_fill = PatternFill("solid", fgColor="F5E08A")
    body_fill_gov = PatternFill("solid", fgColor="D9E6F2")
    body_fill_opp = PatternFill("solid", fgColor="FBF2CC")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Фиксированные ширины колонок, чтобы таблица не растягивалась под длинные имена.
    widths = {1: 10, 2: 10, 3: 18, 4: 18, 5: 18, 6: 18}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row = 2
    for room in (1, 2):
        # Хедер рума.
        ws[f"A{row}"] = "Рум"
        ws[f"B{row}"] = "Позиция"
        ws[f"C{row}"] = "ПРОП"
        ws[f"E{row}"] = "ОПП"

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

        for ref in (f"A{row}", f"B{row}", f"C{row}", f"E{row}"):
            ws[ref].font = bold
            ws[ref].alignment = center
            ws[ref].border = border

        ws[f"C{row}"].fill = gov_fill
        ws[f"E{row}"].fill = opp_fill

        data_row_1 = row + 1
        data_row_2 = row + 2

        # Номер рума — мерджим по вертикали, видим только один раз.
        ws.merge_cells(start_row=data_row_1, start_column=1, end_row=data_row_2, end_column=1)
        ws[f"A{data_row_1}"] = room
        ws[f"A{data_row_1}"].alignment = center
        ws[f"A{data_row_1}"].border = border
        ws[f"A{data_row_1}"].font = bold

        for position, r in ((1, data_row_1), (2, data_row_2)):
            ws[f"B{r}"] = position
            ws[f"B{r}"].alignment = center
            ws[f"B{r}"].border = border

            for col in (3, 4):
                ws.cell(r, col).fill = body_fill_gov
            for col in (5, 6):
                ws.cell(r, col).fill = body_fill_opp

            for col in range(1, 7):
                ws.cell(r, col).alignment = center
                ws.cell(r, col).border = border

            side_names = {
                "ПРОП": assignment.get((room, position, "ПРОП"), []),
                "ОПП": assignment.get((room, position, "ОПП"), []),
            }

            gov = side_names["ПРОП"] + ["", ""]
            ws[f"C{r}"] = gov[0]
            ws[f"D{r}"] = gov[1]

            opp = side_names["ОПП"] + ["", ""]
            ws[f"E{r}"] = opp[0]
            ws[f"F{r}"] = opp[1]

            ws.row_dimensions[r].height = 30

        row += 4  # отступ до следующего рума

    # Форматирование листа участников.
    for cell in ws_p[1]:
        cell.font = bold
        cell.alignment = center
        cell.border = border

    for row_cells in ws_p.iter_rows():
        for cell in row_cells:
            cell.alignment = center
            cell.border = border

    for col_idx, width in {1: 6, 2: 22, 3: 12, 4: 10, 5: 10}.items():
        ws_p.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)