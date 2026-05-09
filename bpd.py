from __future__ import annotations 
import random
import sys
from dataclasses import dataclass
from typing import Dict, List, Optional, Set, Tuple
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MAX_PEOPLE = 16 # 14: глобальные настройки: лимит на 2 рума × 4 позиции × 2 человека = 16
OUTPUT_FILE = "bp_draw.xlsx"
SEED = None  # можно задать высоту для повторяющихся блоков 


@dataclass
class Participant: # 19: Participant хранит всё про человека: уровень, флаг айрона, предпочтение по открывалке
    name: str
    level: str = ""        # старичок/новичок
    opening: int = 0       # айроны на открывающих
    ironman: bool = False  # айроны


@dataclass(frozen=True)
class Slot: # Slot — координаты места в сетке (рум, позиция 1-я или 2-я, ПРОП/ОПП), frozen для использования как ключ
    room: int      # 1/2
    position: int  # 1/2
    side: str      # "ПРОП" или "ОПП"

    @property
    def opening(self) -> bool:
        return self.position == 1


def prompt_choice(prompt: str, valid: Set[str]) -> str: # prompt_choice крутится в цикле, пока юзер не введёт допустимое значение
    valid_lower = {v.lower() for v in valid}
    while True:
        value = input(prompt).strip().lower()
        if value in valid_lower:
            return value 
        print(f"Пожалуйста, введите одно из: {', '.join(sorted(valid_lower))}")


def prompt_int(prompt: str, valid: Optional[Set[int]] = None) -> int:
    while True: # prompt_int — то же самое, но для чисел; valid= None означает "любое целое подойдёт"
        raw = input(prompt).strip()
        try:
            value = int(raw)
        except ValueError:
            print("Please enter a number.")
            continue
        if valid is not None and value not in valid:
            print(f"Please enter one of: {sorted(valid)}")
            continue
        return value


def parse_index_selection(raw: str, max_index: int) -> Set[int]: # parse_index_selection парсит формат "1,3,5-7" в множество индексов; нужно для выбора айронов сразу пачкой
    selected: Set[int] = set()
    raw = raw.strip()
    if not raw:
        return selected

    tokens = raw.replace(" ", "").split(",")
    for token in tokens:
        if not token:
            continue
        if "-" in token:
            left, right = token.split("-", 1)
            start = int(left)
            end = int(right)
            if start > end:
                start, end = end, start
            for n in range(start, end + 1):
                if 1 <= n <= max_index:
                    selected.add(n)
        else:
            n = int(token)
            if 1 <= n <= max_index:
                selected.add(n)
    return selected


def collect_participants() -> List[Participant]:
    print("\nРегистрация участников")
    print("Последовательно введите имена участников.")
    print("Жмякните Enter для пустой строки, или напишите 'done' для завершения реєстрації.\n")

    participants: List[Participant] = []
    while True:
        name = input(f"Участники #{len(participants) + 1}: ").strip()
        if not name or name.lower() == "done":
            break
        participants.append(Participant(name=name))
        if len(participants) >= MAX_PEOPLE:
            print(f"\nReached the maximum of {MAX_PEOPLE} participants.")
            break
    return participants


def choose_ironmen(participants: List[Participant]) -> None: # цикл регистрации останавливается либо по пустой строке, либо по "done", либо по достижению MAX_PEOPLE
    print("\nIronman selection")
    print("Enter participant numbers to mark as ironmen.")
    print("Example: 1,4,7-9")
    print("Press Enter for none.\n")

    while True:
        raw = input("Айроны: ").strip()
        try:
            selected = parse_index_selection(raw, len(participants))
            break
        except ValueError:
            print("Please use only numbers, commas, and ranges like 2-5.")

    for idx in selected:
        participants[idx - 1].ironman = True


def choose_mode() -> int: # choose_mode определяет всю дальнейшую логику расстановки — после этого пути расходятся
    print("\nСтиль расстановки")
    print("1 = Пары старички + новички")
    print("2 = Новички и старички в разных румах")
    return prompt_int("Изберите стиль (1/2): ", valid={1, 2})


def collect_levels_and_opening(participants: List[Participant]) -> None: # открывалку спрашиваем только у айронов, у обычных пар opening всегда 0 (они в открывалке могут оказаться по жребию, не по выбору)
    print("\nИмена участников")
    print("Для каждого участника введите его уровень.")
    print("Открывающая позиция доступна только айронам.\n")

    for i, p in enumerate(participants, start=1):
        print(f"{i}. {p.name}")
        level = prompt_choice("   Уровень [n=новичок, o=старличок]: ", {"n", "o"})
        p.level = "newbie" if level == "n" else "oldman"

        if p.ironman:
            p.opening = prompt_int("   Открывающая позиция [0=нет, 1=да]: ", valid={0, 1})
        else:
            p.opening = 0


def print_summary(participants: List[Participant], mode: int) -> None:
    print("\nФинальная перевірка")
    print(f"Mode: {mode}")
    print("-" * 82)
    print(f"{'#':<4} {'Name':<22} {'Level':<10} {'Ironman':<10} {'Opening':<10}")
    print("-" * 82)
    for i, p in enumerate(participants, start=1):
        print(
            f"{i:<4} {p.name:<22} {p.level:<10} "
            f"{('yes' if p.ironman else 'no'):<10} "
            f"{(p.opening if p.ironman else '-'):<10}"
        )
    print("-" * 82)


def build_slots() -> List[Slot]: #  генерирует все 8 слотов в каноническом порядке (рум 1 - рум 2, позиция 1 - 2, ПРОП - ОПП)
    slots: List[Slot] = []
    for room in (1, 2):
        for position in (1, 2):
            for side in ("ПРОП", "ОПП"):
                slots.append(Slot(room=room, position=position, side=side))
    return slots


def pair_leftovers_randomly(pool: List[Participant]) -> List[List[Participant]]:
    random.shuffle(pool) # сейчас не используется, оставлено на случай рефакторинга
    units: List[List[Participant]] = []
    while len(pool) >= 2:
        units.append([pool.pop(), pool.pop()])
    if pool:
        units.append([pool.pop()])
    return units


def build_units(participants: List[Participant], mode: int) -> List[List[Participant]]: # здесь определяется КТО с КЕМ играет, a НЕ где сидит - слоты распределяются позже
    newbies = [p for p in participants if p.level == "newbie" and not p.ironman]
    oldmen = [p for p in participants if p.level == "oldman" and not p.ironman]

    random.shuffle(newbies)
    random.shuffle(oldmen)

    units: List[List[Participant]] = []

    if mode == 1:
        # Prefer mixed pairs: newbie + oldman
        while newbies and oldmen:
            units.append([newbies.pop(), oldmen.pop()])

        leftovers = newbies + oldmen
        random.shuffle(leftovers)

        while len(leftovers) >= 2:
            units.append([leftovers.pop(), leftovers.pop()])

        if leftovers:
            units.append([leftovers.pop()])

    else:
        # Mode 2: separate by level first
        while len(newbies) >= 2: # режим 1: жадно пихаем смешанные пары новичок+старичок, пока хватает обоих
            units.append([newbies.pop(), newbies.pop()])

        while len(oldmen) >= 2:
            units.append([oldmen.pop(), oldmen.pop()])

        leftovers = newbies + oldmen # остатки одного уровня перемешиваем и ставим в пары случайно — деваться некуда
        random.shuffle(leftovers)

        while len(leftovers) >= 2:
            units.append([leftovers.pop(), leftovers.pop()])

        if leftovers:
            units.append([leftovers.pop()])

    random.shuffle(units) # финальный shuffle перемешивает порядок пар, чтобы порядок их обработки не выдавал, кто пришёл первым
    return units


def assign_ironmen( # assign_ironmen — первый этап раздачи слотов; айроны идут раньше пар, потому что у них есть жёсткие ограничения
    participants: List[Participant],
    slots: List[Slot],
    room_for_level: Optional[Dict[str, int]],
) -> Dict[Tuple[int, int, str], List[str]]:
    assignment: Dict[Tuple[int, int, str], List[str]] = {}
    used_slots: Set[Tuple[int, int, str]] = set()

    ironmen = [p for p in participants if p.ironman]
    random.shuffle(ironmen)

    opening_ironmen = [p for p in ironmen if p.opening == 1] # opening_ironmen хотят только позицию 1 (открывающий блок), non_opening_ironmen — куда угодно
    non_opening_ironmen = [p for p in ironmen if p.opening == 0]

    def pick_slot(p: Participant, candidate_pool: List[Slot]) -> Slot: # pick_slot вынесено в локальную функцию, чтобы не дублировать логику предпочтения рума по уровню
        available = [
            s for s in candidate_pool
            if (s.room, s.position, s.side) not in used_slots
        ]
        if not available:
            raise ValueError(f"Нет доступных слотов для айрона {p.name}")

        if room_for_level is not None: # если room_for_level задан (режим 2), сначала пробуем предпочитаемый рум, при отсутствии слотов — fallback на любой доступный
            preferred_room = room_for_level[p.level]
            preferred = [s for s in available if s.room == preferred_room]
            if preferred:
                return random.choice(preferred)
        return random.choice(available)

    opening_slots = [s for s in slots if s.position == 1]

    # Opening-preference ironmen first
    for p in opening_ironmen: # сначала размещаем айронов с opening=1, иначе они могут не получить открывалку, если non_opening раньше захватят все слоты
        chosen = pick_slot(p, opening_slots)
        assignment[(chosen.room, chosen.position, chosen.side)] = [p.name]
        used_slots.add((chosen.room, chosen.position, chosen.side))

    # Remaining ironmen
    for p in non_opening_ironmen:
        chosen = pick_slot(p, slots)
        assignment[(chosen.room, chosen.position, chosen.side)] = [p.name]
        used_slots.add((chosen.room, chosen.position, chosen.side))

    return assignment


def assign_pairs_to_slots( # assign_pairs_to_slots — второй этап; remaining_slots = всё, что не заняли айроны
    pairs: List[List[Participant]],
    slots: List[Slot],
    current_assignment: Dict[Tuple[int, int, str], List[str]],
    mode: int,
    room_for_level: Optional[Dict[str, int]],
) -> Dict[Tuple[int, int, str], List[str]]:
    remaining_slots = [
        s for s in slots
        if (s.room, s.position, s.side) not in current_assignment
    ]

    if len(pairs) > len(remaining_slots):
        raise ValueError(
            f"Недостаточно позиций для свободных участников. "
            f"Пары: {len(pairs)}, свободные слоты: {len(remaining_slots)}"
        )

    def place(pair: List[Participant], candidates: List[Slot]) -> None:
        chosen = random.choice(candidates)
        current_assignment[(chosen.room, chosen.position, chosen.side)] = [p.name for p in pair]
        remaining_slots.remove(chosen)

    def place_preferring(pair: List[Participant], target_room: int) -> None:
        in_room = [s for s in remaining_slots if s.room == target_room]
        place(pair, in_room if in_room else remaining_slots)

    # режим 1 не различает румы по уровню — ставим пары случайно, баланс уровней уже заложен в составе самих пар
    if mode == 1 or room_for_level is None:
        random.shuffle(pairs)
        for pair in pairs:
            place(pair, remaining_slots)
        return current_assignment

    def pair_level(pair: List[Participant]) -> str:
        levels = {p.level for p in pair}
        if len(levels) == 1:
            return next(iter(levels))
        return "mixed"
    # классификация пар по уровню: oldman, newbie, или mixed (mixed возможна только если в режиме 2 пришёл нечётный остаток)
    oldman_pairs = [p for p in pairs if pair_level(p) == "oldman"]
    newbie_pairs = [p for p in pairs if pair_level(p) == "newbie"]
    mixed_pairs = [p for p in pairs if pair_level(p) == "mixed"]
    random.shuffle(oldman_pairs)
    random.shuffle(newbie_pairs)
    random.shuffle(mixed_pairs)

    # однородные пары идут в свой "родной" рум; place_preferring сам разрулит насыщение, если в родном кончились места
    for pair in oldman_pairs:
        place_preferring(pair, room_for_level["oldman"])
    for pair in newbie_pairs:
        place_preferring(pair, room_for_level["newbie"])

    for pair in mixed_pairs: # смешанные пары обрабатываем последними и ставим в рум, где осталось больше мест — равномерно растягиваем "загрязнение"
        r1 = sum(1 for s in remaining_slots if s.room == 1)
        r2 = sum(1 for s in remaining_slots if s.room == 2)
        if r1 == 0:
            target = 2
        elif r2 == 0:
            target = 1
        else:
            target = 1 if r1 >= r2 else 2
        place_preferring(pair, target)

    return current_assignment

def write_excel( # первый лист — список участников, второй — собственно сетка
    participants: List[Participant],
    assignment: Dict[Tuple[int, int, str], List[str]],
    output_path: str,
) -> None:
    wb = Workbook()

    ws_p = wb.active
    ws_p.title = "Участники"
    ws_p.append(["#", "Имя", "Уровень", "Айрон", "Открывающий"])
    for i, p in enumerate(participants, start=1):
        ws_p.append([i, p.name, p.level, "да" if p.ironman else "нет", p.opening if p.ironman else ""])

    ws = wb.create_sheet("Draw")
    # палитра: синий для ПРОП (правительство), жёлтый для ОПП (оппозиция); тёмные оттенки в шапке, светлые в теле
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    gov_fill = PatternFill("solid", fgColor="5B88B4")
    opp_fill = PatternFill("solid", fgColor="F5E08A")
    body_fill_gov = PatternFill("solid", fgColor="D9E6F2")
    body_fill_opp = PatternFill("solid", fgColor="FBF2CC")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # фиксированные ширины колонок, чтобы таблица не растягивалась под длинные имена
    widths = {1: 10, 2: 10, 3: 18, 4: 18, 5: 18, 6: 18}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row = 2
    for room in (1, 2):
        # Хедер
        ws[f"A{row}"] = "Рум"
        ws[f"B{row}"] = "Позиция"
        ws[f"C{row}"] = "ПРОП"
        ws[f"E{row}"] = "ОПП"

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

        for ref in (f"A{row}", f"B{row}", f"C{row}", f"E{row}"):
            ws[ref].font = bold
            ws[ref].alignment = center
            ws[ref].border = border

        ws[f"C{row}"].fill = gov_fill
        ws[f"E{row}"].fill = opp_fill

        data_row_1 = row + 1
        data_row_2 = row + 2

        ws.merge_cells(start_row=data_row_1, start_column=1, end_row=data_row_2, end_column=1)
        ws[f"A{data_row_1}"] = room
        ws[f"A{data_row_1}"].alignment = center
        ws[f"A{data_row_1}"].border = border
        ws[f"A{data_row_1}"].font = bold

        for position, r in ((1, data_row_1), (2, data_row_2)):
            ws[f"B{r}"] = position
            ws[f"B{r}"].alignment = center
            ws[f"B{r}"].border = border

            for col in (3, 4):
                ws.cell(r, col).fill = body_fill_gov
            for col in (5, 6):
                ws.cell(r, col).fill = body_fill_opp

            for col in range(1, 7):
                ws.cell(r, col).alignment = center
                ws.cell(r, col).border = border

            side_names = {
                "ПРОП": assignment.get((room, position, "ПРОП"), []),
                "ОПП": assignment.get((room, position, "ОПП"), []),
            }

            gov = side_names["ПРОП"] + ["", ""]
            ws[f"C{r}"] = gov[0]
            ws[f"D{r}"] = gov[1]

            opp = side_names["ОПП"] + ["", ""]
            ws[f"E{r}"] = opp[0]
            ws[f"F{r}"] = opp[1]

            ws.row_dimensions[r].height = 30

        row += 4  # форматирование

    # Форматирование участников
    for cell in ws_p[1]:
        cell.font = bold
        cell.alignment = center
        cell.border = border

    for row_cells in ws_p.iter_rows():
        for cell in row_cells:
            cell.alignment = center
            cell.border = border

    for col_idx, width in {1: 6, 2: 22, 3: 12, 4: 10, 5: 10}.items():
        ws_p.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)


def main() -> None:
    if SEED is not None:
        random.seed(SEED)

    print("Генератор сетки для БПД")
    participants = collect_participants()

    if not participants:
        print("Участники не введены. Выход.")
        return

    choose_ironmen(participants)

    ironman_count = sum(1 for p in participants if p.ironman)
    capacity = MAX_PEOPLE - ironman_count
    if len(participants) > capacity:
        print(
            f"\nСлишком много участников.\n"
            f"Участников введено: {len(participants)}\n"
            f"Айронов: {ironman_count}\n"
            f"Макс количество с айронами: {capacity}\n"
        )
        return

    mode = choose_mode()
    collect_levels_and_opening(participants)

    print_summary(participants, mode)
    confirm = prompt_choice("Создаём таблицу из имеющихся данных? [y/n]: ", {"y", "n"})
    if confirm != "y":
        print("Отменено.")
        return

    slots = build_slots()
    if mode == 2:
        oldman_room = random.choice([1, 2])
        room_for_level: Optional[Dict[str, int]] = {
            "oldman": oldman_room,
            "newbie": 3 - oldman_room,
        }
    else:
        room_for_level = None

    assignment = assign_ironmen(participants, slots, room_for_level)
    remaining_people = [p for p in participants if not p.ironman]
    units = build_units(remaining_people, mode)
    assignment = assign_pairs_to_slots(units, slots, assignment, mode, room_for_level)

    write_excel(participants, assignment, OUTPUT_FILE)
    print(f"\nСделано. Сохранено: {OUTPUT_FILE}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nОтменено администратором.")
        sys.exit(1)
